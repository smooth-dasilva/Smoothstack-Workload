import logging
import calendar
import datetime
import inspect

from inspect import currentframe, getframeinfo 
from openpyxl import load_workbook

#start logger in global scope
logging.basicConfig(filename='report.log',filemode='w', level=logging.DEBUG)

#low cohesion, does one thing!
def GetData(filename: str):
    try:
        return load_workbook(filename = filename)
    except FileNotFoundError:
        logging.exception("File %s not found. Program ending..." %filename)



""" Get functions are good candidates for unit testing as it return a single value. I for one didnt implement
    it because I wasn't sure how to mock passing spreadsheet through params. This  additional complexity is a todo.
    
    But more imperatively, there's a general problem with the structure of defining these next few functions.
    It seems unnecessary as they do similar things plus this creates alot of dependency in main.
    Understanding better desing patterns, like the observer pattern or mayve factory patter, which I'm not 
    fully-versed with, and would take more time than necessary for this project, but should do the trick
    in making the structure much neater.     """


def GetVOCRollingCell(sheet, month, year):
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=70):  
        for cell in row: 
            if  isinstance(cell.value, datetime.datetime) and cell.value.year==year and calendar.month_name[cell.value.month].lower()==month:
                return (cell.coordinate)
    logging.error("No corresponding value found in header...")
def GetSummaryRollingCell(sheet, month, year):
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=13, max_col=1):  
        for cell in row: 
            if  isinstance(cell.value, datetime.datetime) and cell.value.year==year and calendar.month_name[cell.value.month].lower()==month:
                return (cell.coordinate)
    logging.error("No corresponding value found in header...")

#break these down to be more reusable. hard-coded values are not ideal
def LogVOCRolling(sheet, cell):
    promoters = sheet["A4"].value
    passives = sheet["A6"].value
    detractors = sheet["A8"].value
    if (sheet[cell[0]+"4"].value)>=200:
        logging.info("%s : Good , Value : %s"  %(str(promoters), sheet[cell[0]+"4"].value))
    else:
        logging.info("%s : Bad , Value : %s" %(str(promoters), str(promoters) %sheet[cell[0]+"4"].value))

    if sheet[cell[0]+"4"].value>=200:
        logging.info("%s : Good, Value : %s" %(str(passives), sheet[cell[0]+"6"].value))
    else:
        logging.info("%s : Bad,  Value : %s" %(str(passives), sheet[cell[0]+"6"].value))

    if sheet[cell[0]+"8"].value>=200:
        logging.info("%s : Bad,  Value : %s" %(str(detractors) , sheet[cell[0]+"8"].value))
    else:
        logging.info("%s : Good, Value :  Value : %s" %(str(detractors), sheet[cell[0]+"8"].value))
def LogSummaryRolling(sheet, cell):
    rowNum =  (cell[1:])
    colChar = cell[0]
    for _ in range(5):
        colChar=chr(ord(colChar) + 1)
        logCell=colChar+(rowNum)
        logging.info( "%s : %s" %( sheet[colChar+"1"].value, sheet[logCell].value))
        


def main():
    
    filename = "datasets/expedia_report_monthly_january_2018.xlsx"
    
    """ calendar months library includes an initial element that is empty.
    this is so the index can match with the actual month index.
    we are not so particular. we show off filters along maps and lamda 
    to get the right months string in a list """

    months = list(map( lambda month : month.lower(), list(filter(lambda month: month!="", calendar.month_name)) ))
    workingMonth = filename.split("_")[3]
    workingYear = int ((filename.split("_")[4])[:4])

    #Get workbook  
    
    # but if the string after the 3rd '_' delimitter split is not in months,
    # throw that error in the log. yell at client to format as expected.
      
    if workingMonth in months:
        workbook = GetData(filename)
    else:
        frameinfo = getframeinfo(currentframe())
        logging.error("File given not formatted properly. Ending program from %s function, line %s" %(currentframe().f_code.co_name, frameinfo.lineno ))
        return 

    # TODO: handle wehn  sheet tab is named properly


    vocSheet=workbook['VOC Rolling MoM']
    summarySheet = workbook['Summary Rolling MoM']

    vocCell = (GetVOCRollingCell(vocSheet, workingMonth, workingYear))

    if vocCell!=None:
        logging.info("Begin VOC log-----------------------")
        LogVOCRolling(vocSheet,vocCell)
        logging.info("End VOC log-----------------------")
    summaryCell= GetSummaryRollingCell(summarySheet, workingMonth, workingYear)
    if summaryCell!=None:
        logging.info("Being summary log-----------------------")
        LogSummaryRolling(summarySheet, summaryCell)
        logging.info("End summary log-----------------------")




#bow down to best practices overloads.
if __name__ == "__main__":
    main()


""" This is solely for practice purposes does not relate to app. 
the code above needs heavy refactoring in order to be used sensibly with unittest.
TDD is difficult. """

class TestUnitTests:
    test: str

    def __init__(self) -> None:
        self.test = "test passed..."
